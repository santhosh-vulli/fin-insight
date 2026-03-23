# app/api/ingest_router.py

import io, os, json, uuid, asyncio, logging, traceback
from datetime import datetime
from typing import AsyncGenerator, Any

from google import genai
from fastapi import APIRouter, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse

from app.database.db import get_connection

router = APIRouter(prefix="/ingest", tags=["ingest"])
log    = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# DEBUG ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/ping")
def ping():
    return {"ok": True, "message": "ingest router is alive"}


@router.get("/diagnose")
def diagnose():
    """Check every dependency. Open http://localhost:8000/ingest/diagnose"""
    result: dict[str, Any] = {}

    api_key = os.environ.get("GEMINI_API_KEY", "")
    result["gemini_api_key"] = {
        "set":    bool(api_key),
        "prefix": api_key[:12] + "..." if len(api_key) > 12 else "(empty)",
    }

    try:
        import openpyxl
        result["openpyxl"] = {"ok": True, "version": openpyxl.__version__}
    except ImportError as e:
        result["openpyxl"] = {"ok": False, "error": str(e)}

    try:
        conn = get_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT 1")
        conn.close()
        result["database"] = {"ok": True}
    except Exception as e:
        result["database"] = {"ok": False, "error": str(e)}

    try:
        from google import genai as _g
        result["google_genai_lib"] = {"ok": True}
    except Exception as e:
        result["google_genai_lib"] = {"ok": False, "error": str(e)}

    if api_key:
        try:
            _client = genai.Client(api_key=api_key)
            _resp = _client.models.generate_content(
                model="gemini-2.5-flash",
                contents="say ok",
            )
            result["gemini_api"] = {"ok": True, "response": _resp.text.strip()[:20]}
        except Exception as e:
            result["gemini_api"] = {"ok": False, "error": str(e)}
    else:
        result["gemini_api"] = {"ok": False, "error": "GEMINI_API_KEY not set"}

    all_ok = all(v.get("ok") for v in result.values() if isinstance(v, dict))
    return JSONResponse({"status": "ok" if all_ok else "issues_found", "checks": result})


@router.get("/test-sse")
async def test_sse():
    """SSE pipe test with mock multi-sheet data."""
    async def _gen():
        yield _sse("phase",    {"phase": "parsing"})
        await asyncio.sleep(0.3)
        yield _sse("phase",    {"phase": "extracting", "model": "mock"})
        await asyncio.sleep(0.5)
        yield _sse("progress", {"pct": 50})
        await asyncio.sleep(0.3)
        yield _sse("progress", {"pct": 100})
        yield _sse("done", {
            "ingest_id": "test-sse-ok",
            "rows": [
                {"account_id": "Revenue",      "period": "2024-01", "amount": 950000, "account_type": "revenue",  "value_type": "amount",     "currency": "USD"},
                {"account_id": "COGS",         "period": "2024-01", "amount": 720000, "account_type": "cogs",     "value_type": "amount",     "currency": "USD"},
                {"account_id": "Gross Profit", "period": "2024-01", "amount": 230000, "account_type": "derived",  "value_type": "amount",     "currency": "USD"},
                {"account_id": "GP %",         "period": "2024-01", "amount": 0.242,  "account_type": "kpi_pct",  "value_type": "percentage", "display": "24.2%", "currency": "USD"},
                {"account_id": "EBITDA",       "period": "2024-01", "amount": 80000,  "account_type": "ebitda",   "value_type": "amount",     "currency": "USD"},
                {"account_id": "EBITDA %",     "period": "2024-01", "amount": 0.084,  "account_type": "kpi_pct",  "value_type": "percentage", "display": "8.4%",  "currency": "USD"},
            ],
            "sheets": {
                "Plan": [
                    {"account_id": "Revenue",      "period": "2024-01", "amount": 950000, "account_type": "revenue", "value_type": "amount"},
                    {"account_id": "COGS",         "period": "2024-01", "amount": 720000, "account_type": "cogs",    "value_type": "amount"},
                    {"account_id": "Gross Profit", "period": "2024-01", "amount": 230000, "account_type": "derived", "value_type": "amount"},
                    {"account_id": "GP %",         "period": "2024-01", "amount": 0.242,  "account_type": "kpi_pct", "value_type": "percentage", "display": "24.2%"},
                    {"account_id": "EBITDA",       "period": "2024-01", "amount": 80000,  "account_type": "ebitda",  "value_type": "amount"},
                    {"account_id": "EBITDA %",     "period": "2024-01", "amount": 0.084,  "account_type": "kpi_pct", "value_type": "percentage", "display": "8.4%"},
                ],
                "KPIs": [],
            },
            "issues": [],
            "summary": {
                "total_rows_extracted": 6,
                "periods_detected": ["2024-01"],
                "currency": "USD",
                "confidence": "high",
                "sheets_populated": ["Plan", "KPIs"],
            },
        })
    return StreamingResponse(_gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@router.post("/test-upload")
async def test_upload(file: UploadFile = File(...)):
    """Parse file without calling AI - confirms file parsing works."""
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(400, "File is empty")

    async def _gen():
        try:
            csv_text = _parse_to_csv(file_bytes, file.filename or "upload.csv")
            line_count = len([l for l in csv_text.splitlines() if l and not l.startswith("#")])
            yield _sse("phase", {"phase": "parsing"})
            await asyncio.sleep(0.1)
            yield _sse("phase", {"phase": "extracting", "model": "mock-no-ai"})
            await asyncio.sleep(0.2)
            yield _sse("progress", {"pct": 100})
            yield _sse("done", {
                "ingest_id": f"test-{uuid.uuid4().hex[:8]}",
                "rows": [
                    {"account_id": f"ParsedRow_{i}", "period": "2024-01",
                     "amount": 1000 * (i + 1), "account_type": "other",
                     "value_type": "amount", "currency": "USD"}
                    for i in range(min(line_count, 5))
                ],
                "sheets": {},
                "issues": [],
                "summary": {
                    "total_rows_extracted": line_count,
                    "periods_detected": ["2024-01"],
                    "currency": "USD",
                    "confidence": "mock",
                    "raw_csv_preview": csv_text[:500],
                },
            })
        except Exception as e:
            yield _sse("error", {"message": f"Parse error: {e}\n{traceback.format_exc()}"})

    return StreamingResponse(_gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ─────────────────────────────────────────────────────────────────────────────
# FILE PARSER
# ─────────────────────────────────────────────────────────────────────────────

def _sse(event: str, data: Any) -> str:
    return f"event: {event}\ndata: {json.dumps(data)}\n\n"


def _parse_to_csv(file_bytes: bytes, filename: str) -> str:
    fname = (filename or "").lower()
    if fname.endswith(".csv"):
        text = file_bytes.decode("utf-8", errors="replace")
        lines = [l for l in text.splitlines() if l.strip()]
        return "\n".join(lines[:201])
    if fname.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(422, "openpyxl not installed. Run: pip install openpyxl")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise HTTPException(422, f"Excel parse error: {e}")
        lines = []
        for ws in wb.worksheets:
            if ws.max_row == 0:
                continue
            lines.append(f"# Sheet: {ws.title}")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 200:
                    break
                if all(v is None for v in row):
                    continue
                lines.append(",".join(
                    "" if v is None else
                    f"{v:.4f}" if isinstance(v, float) and v != int(v) else
                    str(int(v)) if isinstance(v, float) else
                    str(v)
                    for v in row
                ))
        return "\n".join(lines)[:6000]
    raise HTTPException(415, f"Unsupported file type '{filename}'. Upload .csv or .xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# AI PROMPT
# ─────────────────────────────────────────────────────────────────────────────

_SYSTEM = """You are a financial data extraction engine.

Read the uploaded file and extract EVERY row exactly as it appears. Preserve the original order.
Return ONLY valid JSON - no markdown fences, no explanation.

{
  "rows": [
    {
      "account_id": "<exact row label from the file>",
      "period": "YYYY-MM",
      "amount": <number>,
      "value_type": "amount | percentage | ratio"
    }
  ],
  "issues": [{"type": str, "message": str}],
  "summary": {
    "periods_detected": [str],
    "currency": str,
    "confidence": "high | medium | low"
  }
}

EXTRACTION RULES:
1. account_id = the exact label text from the file (e.g. "Revenue", "GP%", "Total Headcount")
2. Preserve the original row order from the file exactly
3. Period mapping: map column headers Jan/Feb/Mar... to YYYY-MM using the year in the file. If no year found use 2024
4. Skip H1/H2/Annual Total/Full Year columns only - these are calculated, not source data
5. Monetary rows: amount = plain number, value_type = "amount"
6. Percentage rows (any row ending in %, like GP%, EBITDA%): amount = decimal (24.2% becomes 0.242), value_type = "percentage"
7. Ratio rows (Revenue/HC, per-unit): amount = raw ratio, value_type = "ratio"
8. Negative amounts are valid
9. Include ALL rows - headcount, assumptions, drivers, KPIs, everything
10. Skip only completely empty rows
"""


def _call_gemini_sync(csv_text: str, api_key: str) -> str:
    client = genai.Client(api_key=api_key)
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=csv_text,
        config={"system_instruction": _SYSTEM},
    )
    return response.text.strip()


def _validate_rows(rows: list, sheet_name: str, issues: list) -> list:
    clean = []
    for i, row in enumerate(rows):
        try:
            row["amount"] = float(row["amount"])
        except (TypeError, ValueError):
            issues.append({"type": "invalid_amount", "message": f"{sheet_name} row {i+1} '{row.get('account_id')}': non-numeric - skipped"})
            continue
        from datetime import datetime as _dt
        period = str(row.get("period", ""))
        if len(period) < 7 or (len(period) > 4 and period[4] != "-"):
            for fmt in ("%Y/%m", "%m/%Y", "%b %Y", "%B %Y"):
                try:
                    row["period"] = _dt.strptime(period, fmt).strftime("%Y-%m")
                    break
                except ValueError:
                    pass
            else:
                issues.append({"type": "date_parse_error", "message": f"{sheet_name} row {i+1}: period '{period}' defaulted to 2024-01"})
                row["period"] = "2024-01"
        row["sheet"] = sheet_name
        clean.append(row)
    return clean


def _persist_sync(rows: list[dict], ingest_id: str) -> None:
    if not rows:
        return
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS fact_financials_staging (
                        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                        ingest_id TEXT NOT NULL,
                        account_id TEXT NOT NULL,
                        cost_center_id TEXT NOT NULL DEFAULT 'default',
                        period TEXT NOT NULL,
                        amount NUMERIC(18,4) NOT NULL,
                        account_type TEXT,
                        currency TEXT DEFAULT 'USD',
                        ingested_at TIMESTAMPTZ DEFAULT NOW(),
                        promoted BOOLEAN DEFAULT FALSE
                    )
                """)
                cur.executemany(
                    "INSERT INTO fact_financials_staging (ingest_id, account_id, cost_center_id, period, amount, account_type, currency) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    [(ingest_id, str(r.get("account_id","Unknown")), str(r.get("cost_center_id","default")), str(r.get("period","2024-01")), float(r.get("amount",0)), str(r.get("account_type","other")), str(r.get("currency","USD"))) for r in rows],
                )
    except Exception as e:
        log.error("Staging persist failed for %s: %s", ingest_id, e)
    finally:
        conn.close()




async def _persist_async(rows: list[dict], ingest_id: str) -> None:
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _persist_sync, rows, ingest_id)


async def _stream_ingest(file_bytes: bytes, filename: str):
    try:
        yield _sse("phase", {"phase": "parsing"})
        await asyncio.sleep(0.05)
        try:
            csv_text = _parse_to_csv(file_bytes, filename)
        except HTTPException as e:
            yield _sse("error", {"message": f"File parse failed: {e.detail}"}); return
        except Exception as e:
            yield _sse("error", {"message": f"File parse error: {e}"}); return

        api_key = os.environ.get("GEMINI_API_KEY", "")
        if not api_key:
            yield _sse("error", {"message": "GEMINI_API_KEY is not set.\nRun: set GEMINI_API_KEY=AIzaSy...\nGet a free key at: https://aistudio.google.com/apikey\nThen restart uvicorn."}); return

        data_rows = sum(1 for l in csv_text.splitlines() if l and not l.startswith("#"))
        yield _sse("phase", {"phase": "extracting", "model": "gemini-2.5-flash", "rows_hint": data_rows})
        await asyncio.sleep(0.05)
        yield _sse("progress", {"pct": 20})
        await asyncio.sleep(0.05)

        try:
            loop = asyncio.get_event_loop()
            raw_json = await loop.run_in_executor(None, _call_gemini_sync, csv_text, api_key)
        except Exception as e:
            msg = str(e).lower()
            if any(k in msg for k in ("api key", "invalid", "400", "permission", "auth", "unauthorized")):
                yield _sse("error", {"message": "GEMINI_API_KEY is invalid or expired. Get a new key at aistudio.google.com/apikey"})
            else:
                yield _sse("error", {"message": f"Gemini API error: {e}"})
            return

        yield _sse("progress", {"pct": 70})
        await asyncio.sleep(0.05)

        cleaned = raw_json.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("```")[1].lstrip("json").strip()
        if not cleaned:
            yield _sse("error", {"message": "Gemini returned an empty response"}); return

        try:
            extraction = json.loads(cleaned)
        except json.JSONDecodeError as e:
            yield _sse("error", {"message": f"Gemini returned invalid JSON: {e}", "raw_preview": cleaned[:300]}); return

        issues  = extraction.get("issues", [])
        summary = extraction.get("summary", {})

        raw_rows = extraction.get("rows", [])
        clean    = _validate_rows(raw_rows, "Plan", issues)

        summary["total_rows_extracted"] = len(clean)
        summary["sheets_populated"]     = ["Plan"]

        ingest_id = str(uuid.uuid4())
        yield _sse("progress", {"pct": 100})
        yield _sse("done", {
            "ingest_id": ingest_id,
            "rows":      clean,
            "sheets":    {"Plan": clean},
            "issues":    issues,
            "summary":   summary,
        })
        asyncio.create_task(_persist_async(clean, ingest_id))

    except Exception as e:
        log.error("Unhandled error in _stream_ingest: %s", traceback.format_exc())
        yield _sse("error", {"message": f"Unexpected server error: {e}"})


@router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """Real upload - SSE stream with Gemini 2.5 Flash extraction."""
    try:
        file_bytes = await file.read()
    except Exception as e:
        raise HTTPException(400, f"Could not read uploaded file: {e}")
    if not file_bytes:
        raise HTTPException(400, "Uploaded file is empty.")
    return StreamingResponse(
        _stream_ingest(file_bytes, file.filename or "upload.csv"),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.post("/promote/{ingest_id}")
def promote_ingest(ingest_id: str, scenario_id: str, version_id: str):
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO fact_financials (tenant_id, scenario_id, version_id, period_id, account_id, cost_center_id, amount)
                    SELECT 'default', %s, %s,
                        COALESCE((SELECT id FROM dim_period WHERE code = s.period LIMIT 1), gen_random_uuid()),
                        s.account_id, s.cost_center_id, s.amount
                    FROM fact_financials_staging s
                    WHERE s.ingest_id = %s AND s.promoted = FALSE
                    ON CONFLICT DO NOTHING
                """, (scenario_id, version_id, ingest_id))
                cur.execute("UPDATE fact_financials_staging SET promoted=TRUE WHERE ingest_id=%s", (ingest_id,))
        return {"promoted": True, "ingest_id": ingest_id}
    finally:
        conn.close()